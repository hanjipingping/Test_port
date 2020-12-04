# _*_coding:utf-8_*_
# @time :2020/11/18 8:07 下午 
# @Author :hanjiping
# Emil :799652949@qq.com
# File :open_pyxl.py
# @software PyCharm

import openpyxl

wb = openpyxl.load_workbook("../Data/m2.xlsx")

# sheet = wb["m1"]
#
# res = sheet.cell(row=2,column =1)
# print(res.value)

# sheet = wb['m2']
sheet = wb['m1']
# res = sheet.cell(row= 1,column= 1,value = "QA")
# wb.save('m2.xlsx')
# res1 = sheet.cell(row =1,column=1)
# print(res1.value)

sheet_list = list(sheet.rows)
print('张博')




print("这块没问题")
# print(sheet_list)
title = [i.value for i in sheet_list[0]]
res = list(filter(None,title))  #去除列表中的不需要的数据
print(res )
title2 = [i.value for i in sheet_list[1]]
print(title2)









